import csv
import io
import json
import os
import re
import urllib.error
import urllib.request
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session

from .models import PriceItem, VendorQuote
from .schemas import PriceImportRow

FIELD_ALIASES: dict[str, list[str]] = {
    "sku": ["sku", "model", "model number", "part", "part no", "part number", "item", "catalog", "型号", "料号", "编号"],
    "name": ["name", "description", "desc", "product", "item description", "equipment", "material", "产品", "描述", "说明", "名称"],
    "category": ["category", "type", "class", "group", "equipment", "material", "labor", "类别", "类型"],
    "brand": ["brand", "manufacturer", "make", "mfg", "品牌", "厂家", "厂牌"],
    "unit": ["unit", "uom", "unit of measure", "单位"],
    "vendor": ["vendor", "supplier", "dealer", "distributor", "供应商", "经销商"],
    "region": ["region", "market", "branch", "location", "区域", "地区"],
    "unit_cost": ["price", "cost", "unit price", "unit cost", "net", "amount", "报价", "价格", "单价", "成本"],
    "lead_time_days": ["lead", "lead time", "days", "eta", "交期", "天数"],
    "quote_date": ["date", "quote date", "effective", "有效期", "日期"],
}

PRICE_TIERS = {"distributor", "contractor", "ecommerce", "manufacturer"}

PRICE_COLUMN_HINTS = [
    "price",
    "cost",
    "net",
    "amount",
    "报价",
    "价格",
    "单价",
    "售价",
    "成本",
    "distributor",
    "dealer",
    "contractor",
    "ecommerce",
    "e commerce",
    "manufacturer",
    "factory",
    "msrp",
    "list",
    "wholesale",
]

PRICE_COLUMN_EXCLUDES = [
    "diff",
    "difference",
    "variance",
    "delta",
    "margin",
    "profit",
    "markup",
    "vs",
    "%",
    "差异",
    "差额",
    "毛利",
    "利润",
]

TIER_HINTS: dict[str, list[str]] = {
    "distributor": ["distributor", "dealer", "supplier", "wholesale", "distribution", "州价格", "经销", "分销", "批发"],
    "contractor": ["contractor", "installer", "trade", "pro price", "承包商", "安装商", "工程商"],
    "ecommerce": ["ecommerce", "e commerce", "e-commerce", "online", "web", "website", "amazon", "shop", "sales price", "retail", "售价", "电商", "线上", "网店"],
    "manufacturer": ["manufacturer", "mfg", "factory", "msrp", "list price", "factory price", "厂家", "工厂", "出厂", "指导价"],
}

CATEGORY_HINTS = {
    "labor": ["labor", "install", "service", "hour", "工时", "人工", "安装"],
    "material": ["refrigerant", "copper", "line set", "filter", "pad", "wire", "材料", "冷媒", "铜管"],
    "equipment": ["condenser", "furnace", "coil", "handler", "heat pump", "mini split", "seer", "ton", "设备", "主机"],
}

PRICE_RE = re.compile(r"(?<![A-Za-z0-9])[$¥￥]?\s*([0-9]{1,3}(?:,[0-9]{3})+|[0-9]+)(?:\.([0-9]{1,4}))?(?![A-Za-z0-9])")
MAX_REASONABLE_PRICE = 500000


@dataclass
class ParsedTable:
    headers: list[str]
    rows: list[list[str]]
    source_label: str | None = None


def normalize(value: Any) -> str:
    return str(value or "").replace("\u00a0", " ").strip()


def compact(value: str) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", " ", value.lower()).strip()


def normalize_price_tier(value: str | None) -> str:
    tier = compact(value or "").replace(" ", "_")
    aliases = {
        "auto": "auto",
        "distribution": "distributor",
        "dealer": "distributor",
        "supplier": "distributor",
        "contract": "contractor",
        "e_commerce": "ecommerce",
        "online": "ecommerce",
        "retail": "ecommerce",
        "mfg": "manufacturer",
        "factory": "manufacturer",
    }
    tier = aliases.get(tier, tier)
    return tier if tier in PRICE_TIERS or tier == "auto" else "distributor"


def infer_price_tier(*values: str, default_tier: str = "auto") -> str:
    text = compact(" ".join(values))
    for tier, hints in TIER_HINTS.items():
        if any(compact(hint) in text for hint in hints):
            return tier
    normalized_default = normalize_price_tier(default_tier)
    return "distributor" if normalized_default == "auto" else normalized_default


def parse_price(value: str) -> float | None:
    text = normalize(value).replace(",", "").replace("$", "").replace("¥", "").replace("￥", "").strip()
    if re.fullmatch(r"[0-9]+(?:\.[0-9]+)?", text):
        number = float(text)
        return round(number, 2) if 0 <= number <= MAX_REASONABLE_PRICE else None

    candidates: list[float] = []
    for match in PRICE_RE.finditer((value or "").replace(" ", "")):
        whole = match.group(1).replace(",", "")
        decimal = match.group(2) or ""
        try:
            number = float(f"{whole}.{decimal}" if decimal else whole)
            if 0 <= number <= MAX_REASONABLE_PRICE:
                candidates.append(round(number, 2))
        except ValueError:
            continue
    return max(candidates) if candidates else None


def parse_int(value: str) -> int | None:
    match = re.search(r"\d+", value or "")
    return int(match.group(0)) if match else None


def parse_date(value: str) -> date:
    text = normalize(value)
    if not text:
        return date.today()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d", "%m-%d-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return date.today()


def detect_category(*values: str) -> str:
    text = compact(" ".join(values))
    for category, hints in CATEGORY_HINTS.items():
        if any(hint in text for hint in hints):
            return category
    return "equipment"


def pad_rows(rows: list[list[str]], width: int) -> list[list[str]]:
    return [row[:width] + [""] * max(width - len(row), 0) for row in rows]


def score_header_row(row: list[str]) -> int:
    score = 0
    for cell in [compact(cell) for cell in row]:
        for aliases in FIELD_ALIASES.values():
            if any(alias in cell for alias in aliases):
                score += 2
        if cell:
            score += 1
    return score


def header_alias_score(row: list[str]) -> int:
    score = 0
    for cell in [compact(cell) for cell in row]:
        for aliases in FIELD_ALIASES.values():
            if any(alias in cell for alias in aliases):
                score += 1
                break
    return score


def choose_header(rows: list[list[str]]) -> ParsedTable:
    rows = [[normalize(cell) for cell in row] for row in rows if any(normalize(cell) for cell in row)]
    if not rows:
        return ParsedTable(headers=[], rows=[])

    header_index, best_score = max(
        [(index, score_header_row(row)) for index, row in enumerate(rows[:20])],
        key=lambda item: item[1],
    )
    if best_score < 3:
        width = max(len(row) for row in rows)
        return ParsedTable(headers=[f"Column {index + 1}" for index in range(width)], rows=pad_rows(rows, width))

    width = max(len(row) for row in rows[header_index:])
    headers = pad_rows([rows[header_index]], width)[0]
    headers = [header or f"Column {index + 1}" for index, header in enumerate(headers)]
    return ParsedTable(headers=headers, rows=pad_rows(rows[header_index + 1 :], width))


def with_source_label(table: ParsedTable, source_label: str) -> ParsedTable:
    table.source_label = source_label
    return table


def split_tables(rows: list[list[str]]) -> list[ParsedTable]:
    rows = [[normalize(cell) for cell in row] for row in rows]
    header_indexes = [
        index
        for index, row in enumerate(rows)
        if any(row) and header_alias_score(row) >= 2
    ]
    if not header_indexes:
        table = choose_header(rows)
        return [table] if table.headers or table.rows else []

    tables: list[ParsedTable] = []
    for table_number, header_index in enumerate(header_indexes, start=1):
        next_header_index = header_indexes[table_number] if table_number < len(header_indexes) else len(rows)
        table = choose_header(rows[header_index:next_header_index])
        if table.headers or table.rows:
            tables.append(with_source_label(table, f"table {table_number}"))
    return tables


def detect_mapping(headers: list[str], rows: list[list[str]]) -> dict[str, int | None]:
    mapping: dict[str, int | None] = {field: None for field in FIELD_ALIASES}
    normalized_headers = [compact(header) for header in headers]

    for field, aliases in FIELD_ALIASES.items():
        alias_tokens = [compact(alias) for alias in aliases]
        best_score = 0
        best_index: int | None = None
        for index, header in enumerate(normalized_headers):
            if field == "unit" and any(token in header for token in ("price", "cost", "价格", "单价", "成本")):
                continue
            score = sum(1 for alias in alias_tokens if alias and alias in header)
            if field == "unit_cost":
                if score:
                    score = (score * 100) + sum(1 for row in rows[:20] if index < len(row) and parse_price(row[index]) is not None)
            if score > best_score:
                best_score = score
                best_index = index
        mapping[field] = best_index if best_score > 0 else None

    if mapping["unit_cost"] is None and rows:
        scores = [
            (sum(1 for row in rows[:40] if index < len(row) and parse_price(row[index]) is not None), index)
            for index in range(len(headers))
        ]
        best_count, best_index = max(scores, default=(0, 0))
        if best_count:
            mapping["unit_cost"] = best_index

    return mapping


def detect_price_columns(headers: list[str], rows: list[list[str]], mapping: dict[str, int | None]) -> list[int]:
    candidates: list[tuple[int, int]] = []
    for index, header in enumerate(headers):
        normalized_header = compact(header)
        if not normalized_header:
            continue
        if any(exclude in normalized_header for exclude in PRICE_COLUMN_EXCLUDES):
            continue
        hint_score = sum(1 for hint in PRICE_COLUMN_HINTS if compact(hint) in normalized_header)
        if not hint_score:
            continue
        numeric_count = sum(1 for row in rows[:40] if index < len(row) and parse_price(row[index]) is not None)
        if numeric_count:
            candidates.append((hint_score * 100 + numeric_count, index))

    if not candidates and mapping["unit_cost"] is not None:
        return [mapping["unit_cost"]]

    ordered_indexes = [index for _, index in sorted(candidates, reverse=True)]
    return list(dict.fromkeys(ordered_indexes))


def dedupe_rows(rows: list[PriceImportRow]) -> list[PriceImportRow]:
    unique_rows: list[PriceImportRow] = []
    seen: set[tuple[str, float | None, str, str, str]] = set()
    for row in rows:
        identity = compact(row.sku or row.name)
        key = (identity, row.unit_cost, compact(row.vendor), compact(row.region), compact(row.price_tier))
        if identity and key in seen:
            continue
        if identity:
            seen.add(key)
        unique_rows.append(row)
    return unique_rows


def value(row: list[str], mapping: dict[str, int | None], field: str) -> str:
    index = mapping.get(field)
    return normalize(row[index]) if index is not None and index < len(row) else ""


def table_to_rows(table: ParsedTable, default_vendor: str, default_region: str, source: str, default_price_tier: str = "auto") -> list[PriceImportRow]:
    mapping = detect_mapping(table.headers, table.rows)
    price_columns = detect_price_columns(table.headers, table.rows, mapping)
    parsed_rows: list[PriceImportRow] = []

    for index, row in enumerate(table.rows, start=2):
        sku = value(row, mapping, "sku")
        name = value(row, mapping, "name") or sku
        brand = value(row, mapping, "brand") or None
        category = value(row, mapping, "category") or detect_category(name, sku, brand or "")
        unit = value(row, mapping, "unit") or "each"
        vendor = value(row, mapping, "vendor") or default_vendor
        region = value(row, mapping, "region") or default_region
        row_price_columns = price_columns or [mapping["unit_cost"]]
        row_prices: list[tuple[str, str, float | None]] = []
        for price_column in row_price_columns:
            if price_column is None or price_column >= len(row):
                continue
            price_header = table.headers[price_column] if price_column < len(table.headers) else "Price"
            price_text = normalize(row[price_column])
            row_prices.append((price_header, price_text, parse_price(price_text)))
        if not row_prices:
            row_prices.append(("Price", " ".join(row), parse_price(" ".join(row))))
        errors = []
        if not name:
            errors.append("Missing item name or model")
        if not vendor:
            errors.append("Missing vendor")

        if sku and name == sku and not re.search(r"\d", sku):
            continue

        for price_header, _price_text, unit_cost in row_prices:
            row_errors = list(errors)
            if unit_cost is None:
                row_errors.append("Missing valid price")
            price_tier = infer_price_tier(price_header, source, default_tier=default_price_tier)
            price_source = f"{source} - {price_header}" if price_header else source
            if name or sku or unit_cost is not None:
                confidence = min(1, 0.35 + (0.2 if sku else 0) + (0.2 if name else 0) + (0.25 if unit_cost is not None else 0) + (0.1 if vendor else 0))
                if price_tier != "distributor":
                    confidence = min(1, confidence + 0.05)
                parsed_rows.append(
                    PriceImportRow(
                        row_number=index,
                        sku=sku or None,
                        name=name,
                        category=category.lower(),
                        brand=brand,
                        unit=unit,
                        vendor=vendor,
                        region=region,
                        price_tier=price_tier,
                        unit_cost=unit_cost,
                        lead_time_days=parse_int(value(row, mapping, "lead_time_days")),
                        quote_date=parse_date(value(row, mapping, "quote_date")),
                        source=price_source,
                        notes=None,
                        confidence=confidence,
                        errors=row_errors,
                    )
                )
    return parsed_rows


def parse_csv_like(content: bytes, delimiter: str) -> ParsedTable:
    text = content.decode("utf-8-sig", errors="replace")
    return choose_header(list(csv.reader(io.StringIO(text), delimiter=delimiter)))


def parse_excel_tables(content: bytes) -> list[ParsedTable]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    tables: list[ParsedTable] = []
    for sheet in workbook.worksheets:
        if sheet.sheet_state != "visible":
            continue
        rows = [[normalize(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
        for table in split_tables(rows):
            source_label = f"{sheet.title} - {table.source_label}" if table.source_label else sheet.title
            tables.append(with_source_label(table, source_label))
    return tables


def parse_pdf_tables(content: bytes) -> list[ParsedTable]:
    import pdfplumber

    tables: list[ParsedTable] = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            rows: list[list[str]] = []
            for table in page.extract_tables() or []:
                rows.extend([[normalize(cell) for cell in row] for row in table])
            text = page.extract_text() or ""
            for line in text.splitlines():
                cells = [part for part in re.split(r"\s{2,}|\t+", line.strip()) if part]
                if len(cells) >= 2 and any(parse_price(cell) is not None for cell in cells):
                    rows.append(cells)
            table = choose_header(rows)
            if table.headers or table.rows:
                tables.append(with_source_label(table, f"page {page.page_number}"))
    return tables


def ai_schema() -> dict[str, Any]:
    return {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "rows": {
                "type": "array",
                "items": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "row_number": {"type": "integer"},
                        "sku": {"type": ["string", "null"]},
                        "name": {"type": "string"},
                        "category": {"type": "string", "enum": ["equipment", "material", "labor", "subcontractor", "other"]},
                        "brand": {"type": ["string", "null"]},
                        "unit": {"type": "string"},
                        "vendor": {"type": "string"},
                        "region": {"type": "string"},
                        "price_tier": {"type": "string", "enum": ["distributor", "contractor", "ecommerce", "manufacturer"]},
                        "unit_cost": {"type": ["number", "null"]},
                        "lead_time_days": {"type": ["integer", "null"]},
                        "quote_date": {"type": "string"},
                        "notes": {"type": ["string", "null"]},
                        "confidence": {"type": "number"},
                        "errors": {"type": "array", "items": {"type": "string"}},
                    },
                    "required": ["row_number", "sku", "name", "category", "brand", "unit", "vendor", "region", "price_tier", "unit_cost", "lead_time_days", "quote_date", "notes", "confidence", "errors"],
                },
            }
        },
        "required": ["rows"],
    }


def ai_prompt(rows: list[PriceImportRow], default_vendor: str, default_region: str, source: str) -> str:
    payload = [row.model_dump(mode="json") for row in rows[:200]]
    return (
        "Normalize HVAC vendor price rows into clean JSON. Keep prices numeric. "
        "Set price_tier to distributor, contractor, ecommerce, or manufacturer based on the price column/header. "
        "Use defaults when missing: vendor=" + default_vendor + ", region=" + default_region + ". "
        "Return only rows that look like price records. Source file: " + source + "\n\n" + json.dumps(payload, ensure_ascii=False)
    )


def rows_from_ai_json(text: str, source: str) -> list[PriceImportRow]:
    data = json.loads(text)
    normalized = []
    for item in data.get("rows", []):
        item["quote_date"] = parse_date(str(item.get("quote_date") or ""))
        item["price_tier"] = normalize_price_tier(item.get("price_tier"))
        normalized.append(PriceImportRow(**item, source=source))
    return normalized


def normalize_with_openai(rows: list[PriceImportRow], default_vendor: str, default_region: str, source: str) -> tuple[list[PriceImportRow], bool]:
    if not os.getenv("OPENAI_API_KEY"):
        return rows, False
    try:
        from openai import OpenAI

        client = OpenAI()
        response = client.responses.create(
            model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
            input=ai_prompt(rows, default_vendor, default_region, source),
            text={"format": {"type": "json_schema", "name": "hvac_price_rows", "schema": ai_schema(), "strict": True}},
        )
        normalized = rows_from_ai_json(response.output_text, source)
        return normalized or rows, bool(normalized)
    except Exception:
        return rows, False


def normalize_with_ollama(rows: list[PriceImportRow], default_vendor: str, default_region: str, source: str) -> tuple[list[PriceImportRow], bool]:
    model = os.getenv("OLLAMA_MODEL")
    if not model:
        return rows, False

    base_url = os.getenv("OLLAMA_BASE_URL", "http://localhost:11434").rstrip("/")
    url = f"{base_url}/generate" if base_url.endswith("/api") else f"{base_url}/api/generate"
    body = {
        "model": model,
        "system": "You clean HVAC price sheets and return strict JSON only.",
        "prompt": ai_prompt(rows, default_vendor, default_region, source),
        "stream": False,
        "format": ai_schema(),
        "options": {"temperature": 0},
    }

    try:
        request = urllib.request.Request(
            url,
            data=json.dumps(body).encode("utf-8"),
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(request, timeout=int(os.getenv("OLLAMA_TIMEOUT", "120"))) as response:
            data = json.loads(response.read().decode("utf-8"))
        normalized = rows_from_ai_json(data.get("response", "{}"), source)
        return normalized or rows, bool(normalized)
    except (OSError, urllib.error.URLError, json.JSONDecodeError, ValueError):
        return rows, False


def normalize_provider(provider: str | None) -> str:
    selected = (provider or os.getenv("AI_PROVIDER", "rules")).strip().lower()
    aliases = {"rule": "rules", "deterministic": "rules", "none": "rules", "local": "ollama"}
    selected = aliases.get(selected, selected)
    return selected if selected in {"rules", "openai", "ollama", "auto"} else "rules"


def ai_normalize_rows(rows: list[PriceImportRow], default_vendor: str, default_region: str, source: str, provider: str | None = None) -> tuple[list[PriceImportRow], bool, str]:
    if not rows:
        return rows, False, normalize_provider(provider)

    selected = normalize_provider(provider)
    if selected == "rules":
        return rows, False, "rules"

    if selected == "openai":
        normalized, used = normalize_with_openai(rows, default_vendor, default_region, source)
        return normalized, used, "openai"

    if selected == "ollama":
        normalized, used = normalize_with_ollama(rows, default_vendor, default_region, source)
        return normalized, used, "ollama"

    normalized, used = normalize_with_openai(rows, default_vendor, default_region, source)
    if used:
        return normalized, True, "openai"
    normalized, used = normalize_with_ollama(rows, default_vendor, default_region, source)
    return normalized, used, "ollama" if used else "rules"


def parse_price_file(filename: str, content: bytes, default_vendor: str = "Imported", default_region: str = "default", ai_provider: str | None = None, default_price_tier: str = "auto") -> tuple[list[PriceImportRow], bool, str]:
    extension = Path(filename).suffix.lower()
    if extension == ".csv":
        tables = [parse_csv_like(content, ",")]
    elif extension == ".tsv":
        tables = [parse_csv_like(content, "\t")]
    elif extension in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        tables = parse_excel_tables(content)
    elif extension == ".pdf":
        tables = parse_pdf_tables(content)
    else:
        raise ValueError("Unsupported file type. Upload .xlsx, .csv, .tsv, or .pdf.")

    rows: list[PriceImportRow] = []
    for table in tables:
        source = f"{filename} - {table.source_label}" if table.source_label else filename
        rows.extend(table_to_rows(table, default_vendor, default_region, source, default_price_tier))
    rows = dedupe_rows(rows)
    return ai_normalize_rows(rows, default_vendor, default_region, filename, ai_provider)


def import_rows(db: Session, rows: list[PriceImportRow]) -> dict[str, int]:
    created_items = 0
    created_quotes = 0
    skipped_rows = 0

    for row in rows:
        if row.errors or row.unit_cost is None:
            skipped_rows += 1
            continue

        item = db.scalar(select(PriceItem).where(PriceItem.sku == row.sku)) if row.sku else None
        if not item:
            item = db.scalar(select(PriceItem).where(PriceItem.name == row.name, PriceItem.category == row.category))
        if not item:
            item = PriceItem(sku=row.sku, name=row.name, category=row.category, brand=row.brand, unit=row.unit, notes=row.notes)
            db.add(item)
            db.flush()
            created_items += 1

        db.add(
            VendorQuote(
                item_id=item.id,
                vendor=row.vendor,
                region=row.region,
                price_tier=row.price_tier,
                unit_cost=row.unit_cost,
                lead_time_days=row.lead_time_days,
                quote_date=row.quote_date,
                source=row.source,
                notes=row.notes,
            )
        )
        created_quotes += 1

    db.commit()
    return {"created_items": created_items, "created_quotes": created_quotes, "skipped_rows": skipped_rows}
